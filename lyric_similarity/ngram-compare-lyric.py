from openpyxl import load_workbook
import argparse, sys, operator

parser = argparse.ArgumentParser()
parser.add_argument('artist', help='artist you wanna compare with specific sentence you will provide')
args = parser.parse_args()

artist = args.artist

base = input('비교를 원하는 문장을 입력하세요: ')

wb = load_workbook('%s.xlsx'%artist)
track_sheet = wb.worksheets[0]
sentence_sheet = wb.worksheets[1]
max_row = sentence_sheet.max_row

sentence_list = []
track_id_list = []

for i in range(2, max_row+1):
    sentence_list.append(sentence_sheet['B%d'%i].value)
    track_id_list.append(sentence_sheet['A%d'%i].value)    

max_row = track_sheet.max_row

track_info_dict = {}

for i in range(2, max_row+1):
    track_info_dict[track_sheet['C%d'%i].value] = track_sheet['B%d'%i].value

# n-gram 유사도 비교
def ngram(s, num):
    res = []
    slen = len(s) - num + 1
    for i in range(slen):
        ss = s[i:i+num]
        res.append(ss)
    return res
def diff_ngram(sa, sb, num):
    a = ngram(sa, num)
    b = ngram(sb, num)
    r = []
    cnt = 0
    for i in a:
        for j in b:
            if i == j:
                cnt += 1
                r.append(i)
    return cnt / len(a), r

two_gram_score_list = []
two_gram_word_list = []
three_gram_score_list = []
three_gram_word_list = []

for s in sentence_list:
    # 2-gram
    r2, word2 = diff_ngram(base, s, 2)
    two_gram_score_list.append(r2)
    two_gram_word_list.append(word2)

    # 3-gram
    r3, word3  = diff_ngram(base, s, 3)
    three_gram_score_list.append(r2)
    three_gram_word_list.append(word2)


print('\n비교 대상 문장: %s\n'%base)
print('유사도 15%를 넘는 문장을 출력합니다.')

# 3-gram
three_max_index = three_gram_score_list.index(max(three_gram_score_list))

three_max_track_id = track_id_list[three_max_index]
print('\n3-gram 분석 결과 가장 유사한 문장: %s'%sentence_list[three_max_index])
print('3-gram 분석 결과 유사한 단어 리스트%s'%three_gram_word_list[three_max_index])
print('3-gram 분석 결과 가장 유사한 곡: %s\n'%track_info_dict[three_max_track_id])

print('유사도 15%가 넘는 모든 곡을 확인합니다.\n')

tmp_list = []
dict_sentence = {}
dict_track = {}

i = 0
for val in three_gram_score_list:
    if val > 0.15:
        if(sentence_list[i] not in tmp_list):      
            tmp_list.append(sentence_list[i])

            dict_sentence[sentence_list[i]] = val
            
            three_gram_track_id = track_id_list[i]
            dict_track[track_info_dict[three_gram_track_id]] = val
    i+=1

dict_sentence = sorted(dict_sentence.items(), key=operator.itemgetter(1), reverse=True)
dict_track = sorted(dict_track.items(), key=operator.itemgetter(1), reverse=True)

for i in range(0, len(dict_sentence)):
    try:
        print('곡: %s / 유사도: %s / 비교된 문장: %s\n'%(dict_track[i][0], dict_sentence[i][1], dict_sentence[i][0]))
    except IndexError:
        sys.exit(1)